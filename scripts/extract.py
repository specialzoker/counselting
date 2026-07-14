import openpyxl, json, sys
from pathlib import Path

_args = [a for a in sys.argv[1:] if a not in ("refs", "hapbul", "special")]
SRC = _args[0] if _args else r"C:\Users\user\Downloads\경기도교육청_2027수시NAVI_수시나비__260707_보호해제.xlsx"
OUT = Path("public/data"); OUT.mkdir(parents=True, exist_ok=True)

def _num(v):
    """숫자 셀만 통과, 아니면(빈칸/'-' 등) None. 타입(number|null)을 정직하게 유지."""
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None

def extract_moojib():
    from openpyxl.utils import column_index_from_string
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["data"]
    rows = list(ws.iter_rows(min_row=2, max_row=6369, values_only=True))
    def idx(letter):
        return column_index_from_string(letter) - 1
    def cut(name_l, p50_l, p70_l, p50_5_l, p70_5_l, r):
        return {"name": r[idx(name_l)], "p50": _num(r[idx(p50_l)]), "p70": _num(r[idx(p70_l)]),
                "p50_5": _num(r[idx(p50_5_l)]), "p70_5": _num(r[idx(p70_5_l)])}
    out = []
    for r in rows:
        if r[idx("D")] is None:  # 대학명 없으면 skip
            continue
        out.append({
            "kwon": r[idx("A")], "region": r[idx("B")], "subRegion": r[idx("C")],
            "univ": r[idx("D")], "moojib26": r[idx("E")], "moojib27": r[idx("H")],
            "gyeyeol": r[idx("I")],
            "gyogwa1": cut("L","M","N","O","P", r),
            "gyogwa2": cut("R","S","T","U","V", r),
            "gyogwa3": cut("X","Y","Z","AA","AB", r),
            "jonghap1": cut("AD","AE","AF","AG","AH", r),
            "jonghap2": cut("AJ","AK","AL","AM","AN", r),
            "code": r[idx("AS")], "banyeong": r[idx("AT")],
            "jeongsiCut70": _num(r[idx("AQ")]), "jeongsiEngHan": r[idx("AR")],
            "studentPercentileCached": _num(r[idx("AU")]),
        })
    json.dump(out, open(OUT/"moojib.json","w",encoding="utf-8"), ensure_ascii=False)
    print("moojib rows:", len(out))

def _formula_str(v):
    """Normalize a cell's formula value to a plain string.
    Array-formula cells return an openpyxl ArrayFormula object (with a .text
    attribute holding the actual formula string), not a str."""
    if v is None:
        return ""
    text = getattr(v, "text", None)
    if text is not None:
        return text
    return str(v)


def extract_calc():
    # Random .cell() access on a read_only sheet is O(n) per call (re-streams),
    # so iterate the (small) 점수계산기 sheet exactly once into row lists instead.
    from openpyxl.utils import column_index_from_string as ci
    wbf = openpyxl.load_workbook(SRC, read_only=True, data_only=False)    # formulas
    wbv = openpyxl.load_workbook(SRC, read_only=True, data_only=True)     # cached values
    wf = wbf["점수계산기"]; wv = wbv["점수계산기"]
    def i(letter): return ci(letter) - 1  # 0-based index into a row tuple

    frows = [list(row) for row in wf.iter_rows(values_only=False)]  # cell objects (formulas)
    vrows = [list(row) for row in wv.iter_rows(values_only=True)]   # cached values

    def fval(row, letter):
        idx = i(letter)
        return _formula_str(row[idx].value) if idx < len(row) else ""
    def vval(row, letter):
        idx = i(letter)
        return row[idx] if idx < len(row) else None

    v2 = vrows[1]  # row 2 = student mock input
    golden = {"studentInput": {
        "kor": vval(v2,"A"), "math": vval(v2,"B"), "tam1": vval(v2,"C"),
        "tam2": vval(v2,"D"), "eng": vval(v2,"E"),
    }, "rows": []}

    patterns = []
    for ridx in range(4, len(vrows)):  # row 5 = index 4
        vrow = vrows[ridx]; frow = frows[ridx]
        code = vval(vrow, "A")
        if code in (None, ""):
            break
        patterns.append({
            "code": code,
            "banyeongText": vval(vrow, "B"),
            "metric": vval(vrow, "C"),
            "weightsRaw": [vval(vrow, c) for c in ["Z","AA","AB","AC"]],
            "subjectFormulas": [fval(frow, c) for c in ["AH","AI","AJ","AK"]],
            "convTable": [vval(vrow, c) for c in ["H","I","J","K","L","M","N","O","P"]],
            "cachedAL": vval(vrow, "AL"),
        })
        golden["rows"].append({"code": code, "al": vval(vrow, "AL")})

    json.dump(patterns, open(OUT/"calc_patterns.json","w",encoding="utf-8"), ensure_ascii=False)
    json.dump(golden, open(OUT/"golden_calc.json","w",encoding="utf-8"), ensure_ascii=False)
    print("calc patterns:", len(patterns))


def extract_search_golden():
    # 검색 결과표(현재 필터 상태의 캐시된 판정 결과)를 종단 검증용 골든으로 추출.
    from openpyxl.utils import column_index_from_string as ci
    wbv = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wbv["검색"]
    def i(letter): return ci(letter) - 1
    rows_all = [list(r) for r in ws.iter_rows(min_row=1, max_row=517, values_only=True)]
    def cell(rownum, letter):
        row = rows_all[rownum - 1]
        idx = i(letter)
        return row[idx] if idx < len(row) else None
    out = []
    for rownum in range(17, 517):
        row = rows_all[rownum - 1]
        univ = row[i("E")] if i("E") < len(row) else None
        if univ in (None, "", "*"):
            continue
        out.append({
            "region": row[i("C")], "univ": univ, "moojib27": row[i("G")],
            "studentP": row[i("P")], "cut70": row[i("Q")], "diff": row[i("S")],
        })
    # 5등급제 토글은 검색!AK5 (판정 수식 IF($AK$5=TRUE,...)의 스위치). V6는 라벨이라 사용하지 않음.
    golden = {"baseMonth": cell(7, "N"), "fiveGrade": bool(cell(5, "AK")), "rows": out}
    json.dump(golden, open(OUT/"golden_search.json","w",encoding="utf-8"), ensure_ascii=False)
    print("search golden rows:", len(out))


def survey_formulas():
    import re, collections
    pats = json.load(open(OUT/"calc_patterns.json",encoding="utf-8"))
    c = collections.Counter()
    for p in pats:
        for f in p["subjectFormulas"]:
            key = re.sub(r"\d+", "N", f)
            c[key] += 1
    with open("scripts/formula_kinds.txt","w",encoding="utf-8") as fh:
        for k,v in c.most_common(): fh.write(f"{v}\t{k}\n")


# ---- 참고 데이터 탭 범용 추출 ----
REF_CONFIGS = [
    # (시트명, 파일키, 헤더행들(1-based), 데이터시작행, 스킵마커행유무)
    ("전형일정", "schedule", [5], 7, True),
    ("전공자율", "jayul", [5], 7, True),
    ("교과반영", "gyogwaBanyeong", [5, 6], 8, True),
    ("특별전형", "special", [6, 7], 9, True),
    ("종합", "jonghap", [6], 8, True),
    ("2028대입", "y2028", [5], 9, True),
    ("수능최저", "suneungMin", [13, 14], 16, True),
    ("논술", "nonsul", [13, 14], 16, True),
]

def _ffill(row):
    """가로 병합셀 대비: None을 왼쪽 값으로 채운다."""
    out, last = [], None
    for v in row:
        if v is not None and str(v).strip() != "":
            last = v
        out.append(last)
    return out

def _header_names(allrows, header_rows, ncol):
    filled = []
    for hr in header_rows:
        raw = allrows[hr - 1]
        row = [raw[c] if c < len(raw) else None for c in range(ncol)]
        filled.append(_ffill(row))
    names = []
    for c in range(ncol):
        parts = []
        for f in filled:
            v = f[c]
            if v is not None and str(v).strip() != "" and str(v) not in parts:
                parts.append(str(v).replace("\n", " ").strip())
        names.append(" · ".join(parts) if parts else f"col{c+1}")
    return names

def extract_refs():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    outdir = OUT / "ref"; outdir.mkdir(parents=True, exist_ok=True)
    index = []
    for sheet, key, header_rows, data_start, skip_marker in REF_CONFIGS:
        ws = wb[sheet]
        allrows = list(ws.iter_rows(values_only=True))
        maxrow = len(allrows)
        # 실제 열 개수: 헤더행 중 가장 넓은 것
        ncol = 0
        for hr in header_rows:
            row = allrows[hr - 1]
            for i, v in enumerate(row):
                if v is not None and str(v).strip() != "":
                    ncol = max(ncol, i + 1)
        names = _header_names(allrows, header_rows, ncol)
        rows = []
        for r in range(data_start, maxrow + 1):
            row = allrows[r - 1]
            cells = [row[c] if c < len(row) else None for c in range(ncol)]
            if all(v is None or str(v).strip() == "" for v in cells):
                continue
            if skip_marker and all((v == "*" or v is None) for v in cells):
                continue
            rows.append([("" if v is None else v) for v in cells])
        # 자동생성 헤더 열(colN) 또는 데이터가 전부 빈 스페이서 열은 제거
        def _auto(n): return n.startswith("col") and n[3:].isdigit()
        keep = [c for c in range(ncol)
                if not _auto(names[c]) and not all(str(row[c]).strip() == "" for row in rows)]
        names = [names[c] for c in keep]
        rows = [[row[c] for c in keep] for row in rows]
        import json as _json
        _json.dump({"sheet": sheet, "columns": names, "rows": rows},
                   open(outdir / f"{key}.json", "w", encoding="utf-8"), ensure_ascii=False)
        index.append({"key": key, "sheet": sheet, "columns": len(names), "rows": len(rows)})
        print(f"ref {sheet}: {len(rows)} rows x {len(names)} cols")
    import json as _json
    _json.dump(index, open(outdir / "index.json", "w", encoding="utf-8"), ensure_ascii=False)



# ---- 합불사례(전형별/모집단위별) 기준대학별 청크 추출 ----
HAPBUL_CONFIGS = [
    ("전형별", "byType"),
    ("모집단위별", "byUnit"),
]

def extract_hapbul():
    import json as _json
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    outdir = OUT / "hapbul"; outdir.mkdir(parents=True, exist_ok=True)
    for sheet, key in HAPBUL_CONFIGS:
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = [str(h) if h is not None else "" for h in next(it)]
        groups = {}
        order = []
        for r in it:
            base = r[0]
            if base is None:
                continue
            if base not in groups:
                groups[base] = []
                order.append(base)
            groups[base].append([("" if v is None else v) for v in r])
        index = []
        for i, base in enumerate(order):
            rows = groups[base]
            _json.dump({"sheet": sheet, "base": base, "columns": header, "rows": rows},
                       open(outdir / f"{key}_{i}.json", "w", encoding="utf-8"), ensure_ascii=False)
            index.append({"id": i, "name": base, "rows": len(rows)})
        _json.dump({"sheet": sheet, "key": key, "columns": header, "bases": index},
                   open(outdir / f"{key}_index.json", "w", encoding="utf-8"), ensure_ascii=False)
        print(f"hapbul {sheet}: {len(order)} bases, {sum(len(g) for g in groups.values())} rows")



# ---- 특수 탭(정적 표) 맞춤 추출 ----
def _special_notice(wb):
    ws = wb["안내필독"]
    intro = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = [str(v).strip() for v in row if v is not None and str(v).strip() != ""]
        if cells:
            intro.append(" ".join(cells))
    return {"sheet": "안내필독", "key": "notice", "intro": intro, "tables": []}


def _special_johgyeon(wb):
    from openpyxl.utils import column_index_from_string as ci
    ws = wb["백분위조견표"]
    note = ws.cell(3, 2).value  # B3: "시트 활용 정보\n1. ...\n..."
    intro = [ln.strip() for ln in str(note).split("\n") if ln.strip()] if note else []

    def clean(v):
        if v is None:
            return ""
        if isinstance(v, str) and v.startswith("#"):  # #VALUE! 등 오류
            return ""
        return v

    def panel(name_c, val_cols, columns, title):
        rows = []
        for r in range(11, 137):
            nm = ws.cell(r, ci(name_c)).value
            if nm is None or str(nm).strip() == "":
                continue
            rows.append([str(nm).strip()] + [clean(ws.cell(r, ci(c)).value) for c in val_cols])
        return {"title": title, "columns": columns, "rows": rows}

    t1 = panel("B", ["O", "P", "Q", "R"], ["대학명", "최대", "중앙값", "최소", "평균"], "대학별 입시결과 범위 (1)")
    t2 = panel("CY", ["DO", "DP", "DQ", "DR"], ["대학명", "최대", "중앙값", "최소", "평균"], "대학별 입시결과 범위 (2)")
    t3 = panel("FE", ["FG", "FH", "FI", "FJ", "FK"],
               ["대학명(세부)", "계열", "70%컷", "정원", "경쟁률", "충원율"], "의약학·교대 세부")
    return {"sheet": "백분위조견표", "key": "johgyeon", "intro": intro, "tables": [t1, t2, t3]}


def _special_grade_conv(wb):
    from openpyxl.utils import column_index_from_string as ci
    ws = wb["등급변환표"]

    def b(r):  # col B 텍스트 한 줄
        v = ws.cell(r, 2).value
        return str(v).strip() if v is not None and str(v).strip() != "" else None

    intro = [t for t in (b(r) for r in list(range(3, 10)) + list(range(18, 27))) if t]

    def cell(r, letter):
        return ws.cell(r, ci(letter)).value

    # 변환 결과(예시 성적): row16을 조합당 한 행으로 세로 전개
    conv_rows = [
        ["전과목", cell(16, "G"), cell(16, "I")],
        ["국수영사과", cell(16, "J"), cell(16, "L")],
        ["국수영과", cell(16, "M"), cell(16, "O")],
        ["국수영사", cell(16, "P"), cell(16, "R")],
    ]
    conv_rows = [[("" if v is None else v) for v in row] for row in conv_rows]
    grade5 = cell(16, "B")
    conv = {"title": f"변환 결과 (예시: 5등급 {grade5})",
            "columns": ["교과조합", "25%-75% 범위", "변환 등급"], "rows": conv_rows}

    # <표1> 5등급→9등급 샘플 (row30~52)
    t1_rows = []
    for r in range(30, 53):
        a, c = cell(r, "B"), cell(r, "F")
        if (a is None or str(a).strip() == "") and (c is None or str(c).strip() == ""):
            continue
        t1_rows.append([("" if a is None else a), ("" if c is None else c)])
    table1 = {"title": "<표1> 5등급→9등급 (샘플)", "columns": ["5등급", "9등급"], "rows": t1_rows}

    # <표2> 석차누적비 샘플
    t2_cols = ["K", "L", "M", "O", "P", "Q"]
    t2_rows = []
    for r in range(30, 53):
        k = cell(r, "K")
        if k is None or str(k).strip() == "":
            continue
        t2_rows.append([("" if cell(r, c) is None else cell(r, c)) for c in t2_cols])
    table2 = {"title": "<표2> 석차누적비 (샘플)",
              "columns": ["5등급", "5등급_석차", "5등급_석차누적비", "9등급", "9등급_석차", "9등급_석차누적비"],
              "rows": t2_rows}
    return {"sheet": "등급변환표", "key": "gradeConv", "intro": intro, "tables": [conv, table1, table2]}


def _special_trend(wb):
    from openpyxl.utils import column_index_from_string as ci
    ws = wb["지원경향"]

    def c8(letter):
        v = ws.cell(8, ci(letter)).value
        return "" if v is None else str(v).strip()

    intro = [f"기준: {c8('C')} / {c8('D')} / {c8('E')} / {c8('G')}"]

    def table(anchor_c, cols, columns, title):
        rows = []
        for r in range(17, 616):  # 헤더 row16, 데이터 row17(순번1)부터
            nm = ws.cell(r, ci(anchor_c)).value
            if nm is None or str(nm).strip() == "":
                continue
            rows.append([("" if ws.cell(r, ci(c)).value is None else ws.cell(r, ci(c)).value) for c in cols])
        return {"title": title, "columns": columns, "rows": rows}

    t1 = table("C", ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"],
               ["순번", "권역", "대학", "전형", "세부전형", "계열", "모집단위", "모집인원", "지원사례수", "합격사례수", "합격률(%)"],
               "모집단위 기준")
    t2 = table("M", ["M", "N", "O", "P", "Q", "R", "S"],
               ["대학", "전형", "세부전형", "계열", "지원사례수", "합격사례수", "합격률(%)"],
               "세부전형 기준")
    return {"sheet": "지원경향", "key": "trend", "intro": intro, "tables": [t1, t2]}


def _special_case_chart(wb):
    # 사례차트: 대학(전형별) 30/50/70%컷 내신 범위. 차트 뷰어(CaseChartTab)가 렌더.
    from openpyxl.utils import column_index_from_string as ci
    ws = wb["사례차트"]
    note = ws.cell(3, 2).value  # B3: "시트 활용 정보\n..."
    intro = [ln.strip() for ln in str(note).split("\n") if ln.strip()] if note else []

    def cell(r, letter):
        return ws.cell(r, ci(letter)).value

    def txt(v):
        return "" if v is None else str(v).strip()

    # 기준 입력행(row6): 권역/전형/계열/기준/학생성적/등급상한/등급하한
    criteria = (f"{txt(cell(6,'B'))} / {txt(cell(6,'C'))} / {txt(cell(6,'D'))} / {txt(cell(6,'E'))}"
                f" · 학생성적 {txt(cell(6,'F'))} · 등급 {txt(cell(6,'G'))}~{txt(cell(6,'H'))}")

    rows = []
    for r in range(9, ws.max_row + 1):
        univ = cell(r, "B")
        if univ is None or str(univ).strip() == "":
            continue
        rows.append({
            "rank": _num(cell(r, "A")),
            "univ": str(univ).strip(),
            "jh": txt(cell(r, "C")),
            "cases": _num(cell(r, "E")),
            "c30": _num(cell(r, "F")),
            "c50": _num(cell(r, "G")),
            "c70": _num(cell(r, "H")),
        })
    return {"sheet": "사례차트", "key": "caseChart", "intro": intro, "criteria": criteria, "rows": rows}


def extract_special():
    import json as _json
    wb = openpyxl.load_workbook(SRC, data_only=True)  # 랜덤 셀 접근 위해 non-read_only
    outdir = OUT / "special"; outdir.mkdir(parents=True, exist_ok=True)
    builders = [_special_notice, _special_johgyeon, _special_grade_conv, _special_trend]
    index = []
    for build in builders:
        d = build(wb)
        _json.dump(d, open(outdir / f"{d['key']}.json", "w", encoding="utf-8"), ensure_ascii=False)
        index.append({"key": d["key"], "sheet": d["sheet"],
                      "tables": len(d["tables"]), "introLines": len(d["intro"])})
        print(f"special {d['sheet']}: {len(d['tables'])} tables, {len(d['intro'])} intro lines")

    # 사례차트: 스키마가 다름(rows/criteria, tables 없음) — 별도 저장.
    dc = _special_case_chart(wb)
    _json.dump(dc, open(outdir / f"{dc['key']}.json", "w", encoding="utf-8"), ensure_ascii=False)
    index.append({"key": dc["key"], "sheet": dc["sheet"], "chartRows": len(dc["rows"]),
                  "introLines": len(dc["intro"])})
    print(f"special {dc['sheet']}: {len(dc['rows'])} chart rows, {len(dc['intro'])} intro lines")

    _json.dump(index, open(outdir / "index.json", "w", encoding="utf-8"), ensure_ascii=False)


if __name__ == "__main__":
    if "refs" in sys.argv:
        extract_refs()
    elif "hapbul" in sys.argv:
        extract_hapbul()
    elif "special" in sys.argv:
        extract_special()
    else:
        extract_moojib()
        extract_calc()
        extract_search_golden()
        survey_formulas()
